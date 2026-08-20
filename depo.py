"""PDGM İş Takip Sistemi için Excel tabanlı veri deposu.

Tasarım hedefi:
- kartlar.xlsx uygulamanın source of truth dosyasıdır.
- Tek Python process + çok thread modeli kullanılır.
- Read/modify/write işlemleri RLock ile korunur.
- Yazmalar temp dosya + os.replace ile yapılır.
- Kritik işlemler öncesi yedek alınır.
- Workflow yalnız dört gerçek durumdan oluşur:
    PLANA ALINDI -> DİZGİDE -> HAZIR -> TESLİM EDİLDİ
- Kaynak Excel'de DURUM boşsa kart saklanır fakat operasyon ekranlarında gösterilmez.

Not: Excel transactional database değildir. Aynı data klasörünü birden fazla Python
process'i paylaşmamalıdır. Bu sınır process lock ile açıkça korunur.
"""

from __future__ import annotations

import atexit
import copy
import os
import re
import shutil
import subprocess
import threading
from datetime import date, datetime
from uuid import uuid4

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel


# ---------------------------------------------------------------------------
# Dosyalar ve workflow sabitleri
# ---------------------------------------------------------------------------

KOK = os.path.dirname(os.path.abspath(__file__))
VERI_KLASORU = os.path.join(KOK, "data")
KARTLAR_DOSYA = os.path.join(VERI_KLASORU, "kartlar.xlsx")
LOG_DOSYA = os.path.join(VERI_KLASORU, "islem_logu.xlsx")
YUKLEME_DOSYA = os.path.join(VERI_KLASORU, "yuklemeler.xlsx")
YEDEK_KLASORU = os.path.join(VERI_KLASORU, "yedekler")
PROCESS_KILIT = os.path.join(VERI_KLASORU, "sunucu.lock")

PLANA_ALINDI = "PLANA ALINDI"
DIZGIDE = "DİZGİDE"
HAZIR = "HAZIR"
TESLIM_EDILDI = "TESLİM EDİLDİ"

GECERLI_DURUMLAR = {PLANA_ALINDI, DIZGIDE, HAZIR, TESLIM_EDILDI}
AKTIF_DURUMLAR = {PLANA_ALINDI, DIZGIDE, HAZIR}

# Template/rapor kodunun daha okunabilir olması için durum sırası tek yerde tutulur.
SIRALAMA = {
    DIZGIDE: 0,
    HAZIR: 1,
    PLANA_ALINDI: 2,
    TESLIM_EDILDI: 3,
    None: 4,
}


# ---------------------------------------------------------------------------
# Exceptions
# ---------------------------------------------------------------------------

class DepoHatasi(Exception):
    """Depo katmanının temel exception sınıfı."""


class KartBulunamadi(DepoHatasi):
    pass


class IsKuralHatasi(DepoHatasi):
    pass


class VeriDogrulamaHatasi(DepoHatasi):
    pass


# ---------------------------------------------------------------------------
# Excel şemaları
# ---------------------------------------------------------------------------

KART_ALANLARI = [
    ("ID", "id"),
    ("Sıra", "sira"),
    ("Talep NO", "talep_no"),
    ("Kart Stok No", "stok_no"),
    ("Talep Sahibi", "talep_sahibi"),
    ("Toplam Adet", "toplam_adet"),
    ("Adet Metni", "adet_metin"),
    ("Plan Haftası", "plan_hafta"),
    ("Plan Başlangıç", "plan_baslama"),
    ("Plan Teslim", "plan_teslim"),
    ("Gerçekleşen Teslim", "gerceklesen_teslim"),
    ("Excel Durumu", "excel_durum"),
    ("PCB", "pcb"),
    ("Durum", "durum"),
    ("Başlangıç Adedi", "baslangic_adet"),
    ("Tamamlanan Adet", "tamamlanan_adet"),
    ("Başlama Zamanı", "baslama_zamani"),
    ("Üretim Bitiş Zamanı", "bitis_zamani"),
    ("Teslim Zamanı", "teslim_zamani"),
    ("Operatör", "operator"),
    ("Not", "aciklama"),
    ("Son Güncelleme", "guncelleme"),
    ("Listede", "aktif"),
    ("Kaynakta Aktif", "source_active"),
    ("Admin Gizli", "admin_gizli"),
    ("Kaynak", "kaynak"),
    ("Anahtar", "anahtar"),
]

LOG_ALANLARI = [
    ("Zaman", "zaman"),
    ("Kullanıcı", "kullanici"),
    ("Rol", "rol"),
    ("İşlem", "islem"),
    ("Talep NO", "talep_no"),
    ("Kart Stok No", "stok_no"),
    ("Adet", "adet"),
    ("Detay", "detay"),
]

YUKLEME_ALANLARI = [
    ("Zaman", "zaman"),
    ("Kullanıcı", "kullanici"),
    ("Dosya", "dosya"),
    ("Okunan Satır", "satir"),
    ("Yeni Kart", "yeni"),
    ("Güncellenen", "guncellenen"),
    ("Kaynakta Olmayan", "pasife_alinan"),
    ("Uyarı", "uyari"),
]

SAYISAL_ALANLAR = {
    "id",
    "sira",
    "toplam_adet",
    "baslangic_adet",
    "tamamlanan_adet",
    "aktif",
    "source_active",
    "admin_gizli",
    "adet",
    "satir",
    "yeni",
    "guncellenen",
    "pasife_alinan",
    "uyari",
}

ZORUNLU_KART_ALANLARI = {
    "id",
    "talep_no",
    "stok_no",
    "toplam_adet",
    "tamamlanan_adet",
    "anahtar",
}


# ---------------------------------------------------------------------------
# In-memory state
# ---------------------------------------------------------------------------

_kilit = threading.RLock()
_kartlar: list[dict] = []
_loglar: list[dict] = []
_yuklemeler: list[dict] = []

LOG_SINIRI = 20_000
LOG_SAKLA = 5_000

BASLIK_DOLGU = PatternFill("solid", fgColor="0F2027")
BASLIK_YAZI = Font(name="Arial", bold=True, color="FFFFFF", size=11)
GOVDE_YAZI = Font(name="Arial", size=10)


# ---------------------------------------------------------------------------
# Genel yardımcılar
# ---------------------------------------------------------------------------

def simdi() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def bugun() -> str:
    return date.today().strftime("%Y-%m-%d")


def _sayi(deger, varsayilan=0) -> int:
    try:
        return int(float(deger))
    except (TypeError, ValueError):
        return varsayilan


def _temiz_metin(deger) -> str:
    return str(deger or "").strip()


def _durum_normalize(deger):
    """Dört gerçek workflow durumunu normalize eder; boş değer None olarak kalır."""
    metin = _temiz_metin(deger)
    if not metin:
        return None

    sade = (
        metin.upper()
        .replace("İ", "I")
        .replace("Ğ", "G")
        .replace("Ü", "U")
        .replace("Ş", "S")
        .replace("Ö", "O")
        .replace("Ç", "C")
    )
    sade = re.sub(r"\s+", " ", sade).strip()

    esleme = {
        "PLANA ALINDI": PLANA_ALINDI,
        "DIZGIDE": DIZGIDE,
        "HAZIR": HAZIR,
        "TESLIM EDILDI": TESLIM_EDILDI,
    }
    return esleme.get(sade)


def tarih_coz(deger):
    """Excel veya kullanıcı girdisini YYYY-MM-DD biçimine çevirir."""
    if deger is None:
        return None

    if isinstance(deger, datetime):
        return deger.strftime("%Y-%m-%d")
    if isinstance(deger, date):
        return deger.strftime("%Y-%m-%d")

    if isinstance(deger, (int, float)) and not isinstance(deger, bool):
        try:
            sonuc = from_excel(deger)
            if isinstance(sonuc, (datetime, date)):
                return sonuc.strftime("%Y-%m-%d")
        except (TypeError, ValueError, OverflowError):
            return None

    metin = str(deger).strip()
    if not metin or metin.upper() in {"-", "YOK", "N/A", "NONE"}:
        return None

    if re.fullmatch(r"\d+(?:\.\d+)?", metin):
        try:
            sonuc = from_excel(float(metin))
            if isinstance(sonuc, (datetime, date)):
                return sonuc.strftime("%Y-%m-%d")
        except (TypeError, ValueError, OverflowError):
            pass

    for kalip in (
        "%Y-%m-%d %H:%M:%S",
        "%d.%m.%Y %H:%M:%S",
        "%d/%m/%Y %H:%M:%S",
        "%d-%m-%Y %H:%M:%S",
        "%Y-%m-%d",
        "%d.%m.%Y",
        "%d/%m/%Y",
        "%Y/%m/%d",
        "%d-%m-%Y",
    ):
        try:
            return datetime.strptime(metin, kalip).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def gun_farki(a, b):
    """a - b gün farkını döndürür."""
    if not a or not b:
        return None
    try:
        t1 = datetime.strptime(str(a)[:10], "%Y-%m-%d").date()
        t2 = datetime.strptime(str(b)[:10], "%Y-%m-%d").date()
    except ValueError:
        return None
    return (t1 - t2).days


def _kart_ref(kart_id: int):
    return next((kart for kart in _kartlar if kart.get("id") == kart_id), None)


def _operasyonda_gorunur_mu(kart: dict) -> bool:
    return (
        kart.get("aktif", 1) == 1
        and kart.get("admin_gizli", 0) != 1
        and kart.get("durum") in GECERLI_DURUMLAR
    )


def _yonetimde_gorunur_mu(kart: dict) -> bool:
    return kart.get("aktif", 1) == 1 and kart.get("admin_gizli", 0) != 1


# ---------------------------------------------------------------------------
# Dosya okuma/yazma
# ---------------------------------------------------------------------------

def _oku(dosya, alanlar, zorunlu_alanlar=frozenset()):
    if not os.path.exists(dosya):
        return []

    try:
        wb = openpyxl.load_workbook(dosya, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise VeriDogrulamaHatasi(
            f"'{os.path.basename(dosya)}' açılamadı: {exc}"
        ) from exc

    try:
        ws = wb[wb.sheetnames[0]]
        satirlar = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not satirlar:
        return []

    basliklar = [str(h or "").strip() for h in satirlar[0]]
    yerlesim = {
        alan: basliklar.index(baslik)
        for baslik, alan in alanlar
        if baslik in basliklar
    }

    eksik = sorted(alan for alan in zorunlu_alanlar if alan not in yerlesim)
    if eksik:
        ters = {alan: baslik for baslik, alan in alanlar}
        eksik_adlar = ", ".join(ters.get(alan, alan) for alan in eksik)
        raise VeriDogrulamaHatasi(
            f"'{os.path.basename(dosya)}' zorunlu sütunları eksik: {eksik_adlar}"
        )

    kayitlar = []
    for satir in satirlar[1:]:
        if not any(hucre not in (None, "") for hucre in satir):
            continue

        kayit = {}
        for _, alan in alanlar:
            index = yerlesim.get(alan)
            deger = satir[index] if index is not None and index < len(satir) else None

            if alan in SAYISAL_ALANLAR:
                kayit[alan] = _sayi(deger, 0) if deger not in (None, "") else None
            elif isinstance(deger, datetime):
                kayit[alan] = deger.strftime("%Y-%m-%d %H:%M:%S")
            else:
                kayit[alan] = str(deger).strip() if deger not in (None, "") else None
        kayitlar.append(kayit)

    return kayitlar

def _excel_hucre_yaz(hucre, deger):
    """'=' ile başlayan kullanıcı metninin Excel formülüne dönüşmesini engeller."""
    hucre.value = deger

    if isinstance(deger, str) and deger.startswith("="):
        hucre.data_type = "s"

def _workbook_uret(alanlar, kayitlar, sayfa_adi):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sayfa_adi
    ws.append([baslik for baslik, _ in alanlar])

    for hucre in ws[1]:
        hucre.fill = BASLIK_DOLGU
        hucre.font = BASLIK_YAZI
        hucre.alignment = Alignment(horizontal="center", vertical="center")

    for kayit in kayitlar:
        satir_no = ws.max_row + 1

        for sutun_no, (_, alan) in enumerate(alanlar, start=1):
            _excel_hucre_yaz(
                ws.cell(row=satir_no, column=sutun_no),
                kayit.get(alan),
            )

    for satir in ws.iter_rows(min_row=2):
        for hucre in satir:
            hucre.font = GOVDE_YAZI

    for sutun, (baslik, alan) in enumerate(alanlar, start=1):
        en = max(
            [len(baslik), 10]
            + [len(str(kayit.get(alan) or "")) for kayit in kayitlar[:300]]
        )
        ws.column_dimensions[get_column_letter(sutun)].width = min(42, en + 3)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    return wb


def _diske_zorla(yol):
    """Temp dosya içeriğinin OS page cache'ten diske inmesini zorlar."""
    try:
        fd = os.open(yol, os.O_RDONLY)
    except OSError:
        return
    try:
        os.fsync(fd)
    except OSError:
        pass
    finally:
        os.close(fd)


def _temp_yaz(hedef, alanlar, kayitlar, sayfa_adi):
    os.makedirs(os.path.dirname(hedef), exist_ok=True)
    temp = f"{hedef}.{uuid4().hex}.yeni"
    wb = _workbook_uret(alanlar, kayitlar, sayfa_adi)
    try:
        wb.save(temp)
    finally:
        wb.close()
    _diske_zorla(temp)
    return temp


def _coklu_yaz(dosyalar):
    """Birden fazla Excel dosyasını temp + replace + rollback yaklaşımıyla yazar."""
    os.makedirs(VERI_KLASORU, exist_ok=True)
    temps = []
    backups = {}
    degisen = []
    commit_basarili = False

    try:
        for hedef, alanlar, kayitlar, sayfa_adi in dosyalar:
            temps.append((hedef, _temp_yaz(hedef, alanlar, kayitlar, sayfa_adi)))

        for hedef, _ in temps:
            if os.path.exists(hedef):
                backup = f"{hedef}.{uuid4().hex}.txn.bak"
                shutil.copy2(hedef, backup)
                backups[hedef] = backup
            else:
                backups[hedef] = None

        for hedef, temp in temps:
            os.replace(temp, hedef)
            degisen.append(hedef)

        commit_basarili = True

    except Exception:
        for hedef in reversed(degisen):
            backup = backups.get(hedef)
            try:
                if backup and os.path.exists(backup):
                    os.replace(backup, hedef)
                    backups[hedef] = None
                elif os.path.exists(hedef):
                    os.remove(hedef)
            except OSError:
                pass
        raise
    finally:
        for _, temp in temps:
            if os.path.exists(temp):
                try:
                    os.remove(temp)
                except OSError:
                    pass
        if commit_basarili:
            for backup in backups.values():
                if backup and os.path.exists(backup):
                    try:
                        os.remove(backup)
                    except OSError:
                        pass


def _yaz(dosya, alanlar, kayitlar, sayfa_adi):
    _coklu_yaz([(dosya, alanlar, kayitlar, sayfa_adi)])


def _gunluk_yedek(dosya):
    if not os.path.exists(dosya):
        return
    os.makedirs(YEDEK_KLASORU, exist_ok=True)
    hedef = os.path.join(
        YEDEK_KLASORU,
        f"{date.today():%Y%m%d}_{os.path.basename(dosya)}",
    )
    if not os.path.exists(hedef):
        shutil.copy2(dosya, hedef)


ANLIK_YEDEK_SAKLA = 30
GUNLUK_YEDEK_GUN = 90
ANLIK_YEDEK_DESEN = re.compile(r"^\d{8}_\d{6}_")


def yedekleri_buda():
    """Yalnız PDGM naming pattern'li yedekleri sınırlar. Hata olursa sessizce geçer."""
    try:
        if not os.path.isdir(YEDEK_KLASORU):
            return

        anlik = []
        for ad in os.listdir(YEDEK_KLASORU):
            yol = os.path.join(YEDEK_KLASORU, ad)

            if os.path.isdir(yol) and ANLIK_YEDEK_DESEN.match(ad):
                try:
                    anlik.append((os.path.getmtime(yol), yol))
                except OSError:
                    pass
                continue

            if re.fullmatch(r"\d{8}_kartlar\.xlsx", ad, flags=re.IGNORECASE):
                try:
                    yas_gun = (
                        datetime.now() - datetime.fromtimestamp(os.path.getmtime(yol))
                    ).days
                    if yas_gun > GUNLUK_YEDEK_GUN:
                        os.remove(yol)
                except OSError:
                    pass

        anlik.sort(reverse=True)
        for _, yol in anlik[ANLIK_YEDEK_SAKLA:]:
            shutil.rmtree(yol, ignore_errors=True)
    except Exception:  # noqa: BLE001
        pass


def anlik_yedek(etiket: str = "once") -> str:
    """Kart/log/yükleme dosyalarının timestamp'li güvenlik kopyasını alır."""
    os.makedirs(YEDEK_KLASORU, exist_ok=True)
    damga = datetime.now().strftime("%Y%m%d_%H%M%S")
    guvenli = re.sub(r"[^0-9A-Za-z_-]+", "_", etiket or "yedek")
    klasor = os.path.join(YEDEK_KLASORU, f"{damga}_{guvenli}")
    os.makedirs(klasor, exist_ok=True)

    for dosya in (KARTLAR_DOSYA, LOG_DOSYA, YUKLEME_DOSYA):
        if os.path.exists(dosya):
            shutil.copy2(dosya, os.path.join(klasor, os.path.basename(dosya)))

    yedekleri_buda()
    return klasor


def yedekleri_getir(adet=12):
    os.makedirs(YEDEK_KLASORU, exist_ok=True)
    sonuc = []

    for ad in os.listdir(YEDEK_KLASORU):
        yol = os.path.join(YEDEK_KLASORU, ad)

        if os.path.isdir(yol):
            kart_dosyasi = os.path.join(yol, os.path.basename(KARTLAR_DOSYA))
            if not os.path.isfile(kart_dosyasi):
                continue
            try:
                mtime = os.path.getmtime(kart_dosyasi)
                boyut = os.path.getsize(kart_dosyasi)
            except OSError:
                continue

            parcalar = ad.split("_", 2)
            etiket = parcalar[2].replace("_", " ") if len(parcalar) >= 3 else "anlık yedek"
            sonuc.append(
                {
                    "ad": ad,
                    "tip": "Anlık",
                    "etiket": etiket,
                    "zaman": datetime.fromtimestamp(mtime).strftime("%d.%m.%Y %H:%M:%S"),
                    "boyut_kb": round(boyut / 1024, 1),
                    "_mtime": mtime,
                }
            )
            continue

        if not os.path.isfile(yol) or not re.fullmatch(
            r"\d{8}_kartlar\.xlsx", ad, flags=re.IGNORECASE
        ):
            continue

        try:
            mtime = os.path.getmtime(yol)
            boyut = os.path.getsize(yol)
        except OSError:
            continue

        sonuc.append(
            {
                "ad": ad,
                "tip": "Günlük",
                "etiket": "günlük otomatik yedek",
                "zaman": datetime.fromtimestamp(mtime).strftime("%d.%m.%Y %H:%M:%S"),
                "boyut_kb": round(boyut / 1024, 1),
                "_mtime": mtime,
            }
        )

    sonuc.sort(key=lambda kayit: kayit["_mtime"], reverse=True)
    for kayit in sonuc:
        kayit.pop("_mtime", None)
    return sonuc if adet is None else sonuc[:adet]


def _yedek_kart_dosyasi_bul(yedek_adi):
    yedek_adi = _temiz_metin(yedek_adi)
    if not yedek_adi:
        raise IsKuralHatasi("Yedek seçilmedi.")
    if os.path.basename(yedek_adi) != yedek_adi:
        raise IsKuralHatasi("Geçersiz yedek adı.")

    yedek_kok = os.path.abspath(YEDEK_KLASORU)
    yol = os.path.abspath(os.path.join(yedek_kok, yedek_adi))
    try:
        if os.path.commonpath([yedek_kok, yol]) != yedek_kok:
            raise IsKuralHatasi("Geçersiz yedek yolu.")
    except ValueError as exc:
        raise IsKuralHatasi("Geçersiz yedek yolu.") from exc

    if os.path.isdir(yol):
        aday = os.path.join(yol, os.path.basename(KARTLAR_DOSYA))
        if os.path.isfile(aday):
            return aday
        raise IsKuralHatasi("Seçilen yedekte kartlar.xlsx bulunamadı.")

    if os.path.isfile(yol) and re.fullmatch(
        r"\d{8}_kartlar\.xlsx", yedek_adi, flags=re.IGNORECASE
    ):
        return yol

    raise IsKuralHatasi("Seçilen yedek kullanılamıyor.")


# ---------------------------------------------------------------------------
# Process lock
# ---------------------------------------------------------------------------



def _process_kilit_sahibi():
    if not os.path.exists(PROCESS_KILIT):
        return None

    try:
        with open(PROCESS_KILIT, encoding="utf-8") as f:
            metin = (f.read() or "").strip()
        ilk = metin.split("|", 1)[0].strip()
        return int(ilk) if ilk else None
    except (OSError, ValueError):
        return None


def _pid_calisiyor_mu(pid: int):
    """PID canlı mı? Karar verilemezse True (fail-closed: kilidi koru)."""
    if pid is None or pid <= 0:
        return False

    if os.name != "nt":
        try:
            os.kill(pid, 0)
            return True
        except ProcessLookupError:
            return False
        except PermissionError:
            return True
        except OSError:
            return True

    try:
        cikti = subprocess.run(
            ["tasklist", "/FI", f"PID eq {pid}", "/NH", "/FO", "CSV"],
            capture_output=True,
            text=True,
            timeout=10,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
    except Exception:
        return True

    if cikti.returncode != 0:
        return True

    return f'"{pid}"' in (cikti.stdout or "")


def _pid_python_mu(pid: int):
    """PID python/pythonw sürecine mi ait? Karar verilemezse True (fail-closed)."""
    if pid is None or pid <= 0:
        return False

    if os.name != "nt":
        try:
            import sys

            if pid == os.getpid():
                return True
            cmdline_yolu = f"/proc/{pid}/cmdline"
            if os.path.exists(cmdline_yolu):
                with open(cmdline_yolu, "rb") as f:
                    cmd = f.read().decode("utf-8", errors="ignore").lower()
                return "python" in cmd
            return True
        except OSError:
            return True

    try:
        cikti = subprocess.run(
            ["tasklist", "/FI", f"PID eq {pid}", "/NH", "/FO", "CSV"],
            capture_output=True,
            text=True,
            timeout=10,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
    except Exception:
        return True

    if cikti.returncode != 0:
        return True

    satir = (cikti.stdout or "").strip().lower()
    if f'"{pid}"' not in satir and str(pid) not in satir:
        return False
    return "python.exe" in satir or "pythonw.exe" in satir


def process_kilidi_al():
    """Aynı data klasörünü ikinci PDGM process'inin açmasını atomik olarak engeller.

    Not:
    - Windows'ta os.kill(pid, 0) yerine tasklist kullanılır.
    - Lock dosyası O_EXCL ile atomik oluşturulur.
    - Stale lock: PID ölüyse veya PID canlı ama python değilse (PID reuse) devralınır.
    - Stale reclaim os.replace ile atomik yapılır (remove+O_EXCL yarışı yok).
    - PID canlı ve python ise reddedilir. tasklist başarısızsa fail-closed.
    """
    os.makedirs(VERI_KLASORU, exist_ok=True)
    mevcut_pid = _process_kilit_sahibi()

    if mevcut_pid == os.getpid():
        return

    if os.path.exists(PROCESS_KILIT):
        canli = mevcut_pid is not None and _pid_calisiyor_mu(mevcut_pid)
        python_sureci = canli and _pid_python_mu(mevcut_pid)

        if canli and python_sureci:
            raise RuntimeError(
                f"data/ klasörü başka bir PDGM process tarafından kilitli "
                f"(PID {mevcut_pid}) ve bu process ŞU AN ÇALIŞIYOR. "
                "İkinci sunucu açmayın."
            )

        sahip = mevcut_pid if mevcut_pid is not None else "bilinmiyor"
        if canli and not python_sureci:
            print(
                f"UYARI: data/sunucu.lock PID {sahip} başka bir uygulamaya ait "
                "(PID reuse). Kalıntı kilit devralınıyor."
            )
        else:
            print(
                f"UYARI: data/sunucu.lock artık çalışmayan bir process'e ait "
                f"(PID {sahip}). Kalıntı kilit devralınıyor."
            )

        # Atomik reclaim: remove+O_EXCL yarışını önlemek için stale lock'u
        # benzersiz bir isme taşı. İki process aynı anda denerse yalnız biri
        # os.replace kazanır; diğeri FileNotFoundError alır ve O_EXCL'de kaybeder.
        stale_yol = f"{PROCESS_KILIT}.stale.{uuid4().hex}"
        try:
            os.replace(PROCESS_KILIT, stale_yol)
        except FileNotFoundError:
            stale_yol = None
        except OSError as exc:
            raise RuntimeError(
                "data/sunucu.lock devralınamadı: "
                f"{exc}. Dosyayı manuel silip tekrar deneyin."
            ) from exc
    else:
        stale_yol = None

    bayraklar = os.O_WRONLY | os.O_CREAT | os.O_EXCL

    try:
        fd = os.open(PROCESS_KILIT, bayraklar)
    except FileExistsError as exc:
        if stale_yol:
            try:
                os.remove(stale_yol)
            except OSError:
                pass
        sahip = _process_kilit_sahibi()
        raise RuntimeError(
            f"data/ klasörü başka bir PDGM process tarafından kilitlendi "
            f"(PID {sahip if sahip is not None else 'bilinmiyor'})."
        ) from exc

    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            f.write(f"{os.getpid()}|{datetime.now():%Y-%m-%d %H:%M:%S}")
            f.flush()
            os.fsync(f.fileno())
    except Exception:
        try:
            os.remove(PROCESS_KILIT)
        except OSError:
            pass
        if stale_yol:
            try:
                os.remove(stale_yol)
            except OSError:
                pass
        raise

    if stale_yol:
        try:
            os.remove(stale_yol)
        except OSError:
            pass

    if getattr(process_kilidi_al, "_atexit_bagli", False):
        return

    def _birak():
        try:
            if _process_kilit_sahibi() == os.getpid():
                os.remove(PROCESS_KILIT)
        except OSError:
            pass

    atexit.register(_birak)
    process_kilidi_al._atexit_bagli = True


# ---------------------------------------------------------------------------
# Kart normalizasyonu ve doğrulama
# ---------------------------------------------------------------------------

def _kart_normalize(kart: dict) -> dict:
    kart = dict(kart)
    kart["id"] = _sayi(kart.get("id"), 0)
    kart["sira"] = _sayi(kart.get("sira"), 0) or None
    kart["toplam_adet"] = _sayi(kart.get("toplam_adet"), 1) or 1
    kart["baslangic_adet"] = _sayi(kart.get("baslangic_adet"), 0)
    kart["tamamlanan_adet"] = _sayi(kart.get("tamamlanan_adet"), 0)
    kart["aktif"] = 0 if kart.get("aktif") == 0 else 1
    kart["source_active"] = 0 if kart.get("source_active") == 0 else 1
    kart["admin_gizli"] = 1 if kart.get("admin_gizli") == 1 else 0
    kart["kaynak"] = _temiz_metin(kart.get("kaynak")) or "EXCEL"
    kart["durum"] = _durum_normalize(kart.get("durum"))

    for alan in ("plan_baslama", "plan_teslim", "gerceklesen_teslim"):
        ham = kart.get(alan)
        if ham in (None, ""):
            kart[alan] = None
            continue
        cozulmus = tarih_coz(ham)
        if not cozulmus:
            raise VeriDogrulamaHatasi(
                f"Kart {kart.get('id') or '?'}: {alan} geçerli bir tarih değil: {ham!r}"
            )
        kart[alan] = cozulmus

    return kart


def _kart_dogrula(kart: dict):
    kart_id = _sayi(kart.get("id"), -1)
    toplam = _sayi(kart.get("toplam_adet"), 0)
    tamam = _sayi(kart.get("tamamlanan_adet"), 0)
    durum = kart.get("durum")

    if kart_id < 1:
        raise VeriDogrulamaHatasi("Kart ID pozitif tam sayı olmalı.")
    if not _temiz_metin(kart.get("talep_no")):
        raise VeriDogrulamaHatasi(f"Kart {kart_id}: Talep NO boş olamaz.")
    if not _temiz_metin(kart.get("stok_no")):
        raise VeriDogrulamaHatasi(f"Kart {kart_id}: Kart Stok No boş olamaz.")
    if toplam < 1:
        raise VeriDogrulamaHatasi(f"Kart {kart_id}: Toplam Adet en az 1 olmalı.")
    if tamam < 0 or tamam > toplam:
        raise VeriDogrulamaHatasi(
            f"Kart {kart_id}: Tamamlanan Adet ({tamam}) 0 ile Toplam Adet ({toplam}) arasında olmalı."
        )
    if durum is not None and durum not in GECERLI_DURUMLAR:
        raise VeriDogrulamaHatasi(
            f"Kart {kart_id}: Geçersiz durum '{durum}'. "
            f"Geçerli durumlar: {', '.join(sorted(GECERLI_DURUMLAR))}."
        )
    if durum in (None, PLANA_ALINDI) and tamam != 0:
        etiket = "Durumu boş" if durum is None else PLANA_ALINDI
        raise VeriDogrulamaHatasi(
            f"Kart {kart_id}: {etiket} kartta Tamamlanan Adet 0 olmalı."
        )
    if durum in (HAZIR, TESLIM_EDILDI) and tamam != toplam:
        raise VeriDogrulamaHatasi(
            f"Kart {kart_id}: {durum} durumunda Tamamlanan Adet Toplam Adet'e eşit olmalı."
        )
    if durum == TESLIM_EDILDI and not kart.get("gerceklesen_teslim"):
        raise VeriDogrulamaHatasi(
            f"Kart {kart_id}: TESLİM EDİLDİ durumunda Gerçekleşen Teslim tarihi zorunlu."
        )
    if kart.get("plan_baslama") and kart.get("plan_teslim"):
        if kart["plan_baslama"] > kart["plan_teslim"]:
            raise VeriDogrulamaHatasi(
                f"Kart {kart_id}: Plan başlangıç tarihi plan teslim tarihinden sonra olamaz."
            )
    if not _temiz_metin(kart.get("anahtar")):
        raise VeriDogrulamaHatasi(f"Kart {kart_id}: Anahtar boş olamaz.")


def _kart_listesi_dogrula(kartlar):
    idler = set()
    anahtarlar = set()

    for kart in kartlar:
        _kart_dogrula(kart)
        if kart["id"] in idler:
            raise VeriDogrulamaHatasi(f"Tekrarlanan kart ID: {kart['id']}")
        if kart["anahtar"] in anahtarlar:
            raise VeriDogrulamaHatasi(f"Tekrarlanan kart anahtarı: {kart['anahtar']}")
        idler.add(kart["id"])
        anahtarlar.add(kart["anahtar"])


# ---------------------------------------------------------------------------
# Başlangıç ve reload
# ---------------------------------------------------------------------------

def _baslatma_uyarisi_yaz(mesaj: str) -> None:
    """Açılış uyarılarını konsola basar ve data/BASLATMA_HATASI.txt'ye ekler."""
    print(mesaj)
    yol = os.path.join(VERI_KLASORU, "BASLATMA_HATASI.txt")
    try:
        with open(yol, "a", encoding="utf-8") as f:
            f.write(f"{datetime.now():%Y-%m-%d %H:%M:%S}  {mesaj}\n")
    except OSError:
        pass


def _bozuk_dosyayi_kenara_al(dosya, alanlar):
    """Okunamayan yardımcı dosyayı yeniden adlandırıp boş liste döner.

    Yalnız islem_logu / yuklemeler için kullanılır. Kart verisi bu yolla sıfırlanmaz.
    """
    try:
        return _oku(dosya, alanlar)
    except VeriDogrulamaHatasi as exc:
        if os.path.exists(dosya):
            damga = datetime.now().strftime("%Y%m%d_%H%M%S")
            bozuk = f"{dosya}.bozuk_{damga}"
            try:
                os.replace(dosya, bozuk)
            except OSError:
                bozuk = dosya
            _baslatma_uyarisi_yaz(
                f"UYARI: '{os.path.basename(dosya)}' okunamadı ({exc}). "
                f"Dosya '{os.path.basename(bozuk)}' olarak kenara alındı; "
                "boş liste ile devam ediliyor. Kart verisi etkilenmedi."
            )
        else:
            _baslatma_uyarisi_yaz(
                f"UYARI: '{os.path.basename(dosya)}' okunamadı ({exc}). "
                "Boş liste ile devam ediliyor."
            )
        return []


def kur():
    global _kartlar, _loglar, _yuklemeler

    with _kilit:
        os.makedirs(VERI_KLASORU, exist_ok=True)

        kartlar = _oku(
            KARTLAR_DOSYA,
            KART_ALANLARI,
            ZORUNLU_KART_ALANLARI if os.path.exists(KARTLAR_DOSYA) else frozenset(),
        )
        kartlar = [_kart_normalize(kart) for kart in kartlar]
        if kartlar:
            _kart_listesi_dogrula(kartlar)

        _kartlar = kartlar
        _loglar = _bozuk_dosyayi_kenara_al(LOG_DOSYA, LOG_ALANLARI)
        _yuklemeler = _bozuk_dosyayi_kenara_al(YUKLEME_DOSYA, YUKLEME_ALANLARI)

        eksikler = []
        if not os.path.exists(KARTLAR_DOSYA):
            eksikler.append((KARTLAR_DOSYA, KART_ALANLARI, _kartlar, "Kartlar"))
        if not os.path.exists(LOG_DOSYA):
            eksikler.append((LOG_DOSYA, LOG_ALANLARI, _loglar, "İşlem Logu"))
        if not os.path.exists(YUKLEME_DOSYA):
            eksikler.append((YUKLEME_DOSYA, YUKLEME_ALANLARI, _yuklemeler, "Yüklemeler"))
        if eksikler:
            _coklu_yaz(eksikler)


def kartlari_diskten_yeniden_yukle():
    """Manuel Excel müdahalesinden sonra kartlar.xlsx'i validate ederek tekrar yükler."""
    global _kartlar

    with _kilit:
        yeni = _oku(KARTLAR_DOSYA, KART_ALANLARI, ZORUNLU_KART_ALANLARI)
        yeni = [_kart_normalize(kart) for kart in yeni]
        _kart_listesi_dogrula(yeni)
        _kartlar = yeni
        return len(_kartlar)


def yeniden_yukle():
    return kartlari_diskten_yeniden_yukle()


def _kartlari_kaydet():
    _gunluk_yedek(KARTLAR_DOSYA)
    _yaz(KARTLAR_DOSYA, KART_ALANLARI, _kartlar, "Kartlar")


# ---------------------------------------------------------------------------
# Log ve yükleme geçmişi
# ---------------------------------------------------------------------------

def _log_kaydi(kullanici, rol, islem, talep_no="", stok_no="", adet=None, detay=""):
    return {
        "zaman": simdi(),
        "kullanici": kullanici,
        "rol": rol,
        "islem": islem,
        "talep_no": talep_no,
        "stok_no": stok_no,
        "adet": adet,
        "detay": detay,
    }


def _log_arsivle_gerekirse():
    global _loglar

    if len(_loglar) <= LOG_SINIRI:
        return

    os.makedirs(YEDEK_KLASORU, exist_ok=True)
    arsivlenecek = _loglar[:-LOG_SAKLA]
    arsiv = os.path.join(
        YEDEK_KLASORU,
        f"{datetime.now():%Y%m%d_%H%M%S}_islem_logu_arsiv.xlsx",
    )
    _yaz(arsiv, LOG_ALANLARI, arsivlenecek, "İşlem Logu")
    _loglar = _loglar[-LOG_SAKLA:]


def log_ekle(kullanici, rol, islem, talep_no="", stok_no="", adet=None, detay=""):
    global _loglar

    with _kilit:
        eski = copy.deepcopy(_loglar)
        try:
            _loglar.append(
                _log_kaydi(kullanici, rol, islem, talep_no, stok_no, adet, detay)
            )
            _log_arsivle_gerekirse()
            _yaz(LOG_DOSYA, LOG_ALANLARI, _loglar, "İşlem Logu")
        except Exception:
            _loglar = eski
            raise


def loglari_getir(adet=None):
    with _kilit:
        secim = list(reversed(_loglar))
        if adet:
            secim = secim[:adet]
        return copy.deepcopy(secim)


def yuklemeleri_getir(adet=None):
    with _kilit:
        secim = list(reversed(_yuklemeler))
        if adet:
            secim = secim[:adet]
        return copy.deepcopy(secim)


# ---------------------------------------------------------------------------
# Görünüm hesapları
# ---------------------------------------------------------------------------

def durum_bilgisi(kart):
    durum = kart.get("durum")
    plan_baslama = kart.get("plan_baslama")
    plan_teslim = kart.get("plan_teslim")
    bugun_iso = bugun()

    bilgi = {
        "rozet": durum or "DURUMU EKSİK",
        "renk": "notr",
        "sapma": None,
        "kalan": None,
        "zaman_yuzde": 0,
        "plan_gun": gun_farki(plan_teslim, plan_baslama),
    }

    if durum is None:
        bilgi["rozet"] = "DURUMU EKSİK"
        bilgi["renk"] = "uyari"
        return bilgi

    if durum == TESLIM_EDILDI:
        teslim = kart.get("gerceklesen_teslim") or str(kart.get("teslim_zamani") or "")[:10]
        sapma = gun_farki(teslim, plan_teslim)
        bilgi["sapma"] = sapma
        bilgi["zaman_yuzde"] = 100
        if sapma is None:
            bilgi["rozet"], bilgi["renk"] = "TESLİM EDİLDİ", "iyi"
        elif sapma > 0:
            bilgi["rozet"], bilgi["renk"] = f"GEÇ TESLİM (+{sapma} gün)", "kotu"
        else:
            bilgi["rozet"], bilgi["renk"] = "ZAMANINDA TESLİM", "iyi"
        return bilgi

    if durum == HAZIR:
        gecikme = gun_farki(bugun_iso, plan_teslim)
        bilgi["zaman_yuzde"] = 100
        if plan_teslim and gecikme is not None and gecikme > 0:
            bilgi["sapma"] = gecikme
            bilgi["rozet"], bilgi["renk"] = f"HAZIR · TESLİM BEKLİYOR (+{gecikme} gün)", "kotu"
        else:
            bilgi["rozet"], bilgi["renk"] = "HAZIR · TESLİM BEKLİYOR", "iyi"
        return bilgi

    if durum == DIZGIDE:
        if kart.get("tamamlanan_adet", 0) >= kart.get("toplam_adet", 1):
            bilgi["rozet"], bilgi["renk"] = "ÜRETİM BİTTİ · HAZIRA ALIN", "uyari"
            bilgi["zaman_yuzde"] = 100
            return bilgi

        kalan = gun_farki(plan_teslim, bugun_iso)
        bilgi["kalan"] = kalan
        baslangic = str(kart.get("baslama_zamani") or plan_baslama or bugun_iso)[:10]
        gecen = gun_farki(bugun_iso, baslangic) or 0
        plan_gun = bilgi["plan_gun"]

        if plan_gun and plan_gun > 0:
            bilgi["zaman_yuzde"] = max(0, min(140, round(gecen / plan_gun * 100)))
        else:
            bilgi["zaman_yuzde"] = 100 if kalan is not None and kalan < 0 else 50

        if kalan is None:
            bilgi["rozet"], bilgi["renk"] = "DİZGİDE", "uyari"
        elif kalan < 0:
            bilgi["sapma"] = -kalan
            bilgi["rozet"], bilgi["renk"] = f"SÜRE AŞILDI ({-kalan} gün)", "kotu"
        elif kalan <= 1:
            bilgi["rozet"] = "SON GÜN" if kalan == 0 else "SON 1 GÜN"
            bilgi["renk"] = "uyari"
        else:
            bilgi["rozet"], bilgi["renk"] = f"PLANINDA ({kalan} gün var)", "iyi"
        return bilgi

    gecikme = gun_farki(bugun_iso, plan_baslama)
    if plan_baslama and gecikme is not None and gecikme > 0:
        bilgi["rozet"], bilgi["renk"], bilgi["sapma"] = (
            f"BAŞLAMADI (+{gecikme} gün)",
            "kotu",
            gecikme,
        )
    elif plan_baslama and gecikme == 0:
        bilgi["rozet"], bilgi["renk"] = "BUGÜN BAŞLAMALI", "uyari"
    else:
        bilgi["rozet"], bilgi["renk"] = PLANA_ALINDI, "notr"
    return bilgi


def kart_gorunumu(kart):
    d = copy.deepcopy(kart)
    d.update(durum_bilgisi(kart))
    d["toplam_adet"] = d.get("toplam_adet") or 1
    d["tamamlanan_adet"] = d.get("tamamlanan_adet") or 0
    d["baslangic_adet"] = d.get("baslangic_adet") or 0
    d["kalan_adet"] = max(0, d["toplam_adet"] - d["tamamlanan_adet"])
    d["adet_yuzde"] = min(100, round(d["tamamlanan_adet"] / d["toplam_adet"] * 100))
    d["gorunur"] = _operasyonda_gorunur_mu(d)
    d["kaynakta_yok"] = d.get("source_active", 1) != 1
    d["is_durumu"] = d.get("durum") or "DURUMU EKSİK"
    d["kaynak_durumu"] = _temiz_metin(d.get("excel_durum"))
    return d


def _kartlari_sirala(kartlar):
    kartlar.sort(
        key=lambda kart: (
            SIRALAMA.get(kart.get("durum"), 9),
            kart.get("plan_baslama") or "9999-12-31",
            kart.get("sira") or 999999,
            kart.get("id") or 0,
        )
    )
    return kartlar


def kartlari_getir(sadece_gorunen=True):
    with _kilit:
        secim = [
            kart for kart in _kartlar
            if not sadece_gorunen or _operasyonda_gorunur_mu(kart)
        ]
        kartlar = [kart_gorunumu(kart) for kart in secim]
    return _kartlari_sirala(kartlar)


def kartlari_yonetim_getir():
    """Admin tablosu için, durumu boş kartlar dahil, gizlenmemiş tüm aktif kartlar."""
    with _kilit:
        kartlar = [
            kart_gorunumu(kart)
            for kart in _kartlar
            if _yonetimde_gorunur_mu(kart)
        ]
    return _kartlari_sirala(kartlar)


def durumu_eksik_kartlari_getir():
    with _kilit:
        kartlar = [
            kart_gorunumu(kart)
            for kart in _kartlar
            if _yonetimde_gorunur_mu(kart) and not kart.get("durum")
        ]
    return _kartlari_sirala(kartlar)


def gizlenen_kartlari_getir():
    with _kilit:
        kartlar = [
            kart_gorunumu(kart)
            for kart in _kartlar
            if kart.get("aktif", 1) == 1 and kart.get("admin_gizli", 0) == 1
        ]
    kartlar.sort(key=lambda kart: (kart.get("guncelleme") or "", kart.get("id") or 0), reverse=True)
    return kartlar


def kart_getir(kart_id):
    kart_id = _sayi(kart_id, -1)
    with _kilit:
        kart = _kart_ref(kart_id)
        return kart_gorunumu(kart) if kart else None


def kart_bul(anahtar):
    with _kilit:
        for kart in _kartlar:
            if kart.get("anahtar") == anahtar:
                return copy.deepcopy(kart)
    return None


def yeni_kimlik():
    with _kilit:
        return max([_sayi(kart.get("id"), 0) for kart in _kartlar] or [0]) + 1


# ---------------------------------------------------------------------------
# Kart + log commit yardımcıları
# ---------------------------------------------------------------------------

def _kart_log_commit():
    _gunluk_yedek(KARTLAR_DOSYA)
    _coklu_yaz(
        [
            (KARTLAR_DOSYA, KART_ALANLARI, _kartlar, "Kartlar"),
            (LOG_DOSYA, LOG_ALANLARI, _loglar, "İşlem Logu"),
        ]
    )


def _atomik_kart_islemi(islem):
    """RAM state rollback kalıbını tek yerde tutar.

    Kartlar in-place mutasyon gördüğü için deepcopy zorunlu.
    Loglara yalnız append yapıldığı için uzunluk + del yeterlidir.
    """
    global _kartlar

    eski_kartlar = copy.deepcopy(_kartlar)
    log_sayisi = len(_loglar)
    try:
        sonuc = islem()
        _kart_listesi_dogrula(_kartlar)
        _kart_log_commit()
        return sonuc
    except Exception:
        _kartlar = eski_kartlar
        del _loglar[log_sayisi:]
        raise


# ---------------------------------------------------------------------------
# Operatör workflow işlemleri
# ---------------------------------------------------------------------------

def kart_baslat(kart_id, adet, kullanici, rol, aciklama=""):
    kart_id = _sayi(kart_id, -1)

    with _kilit:
        kart = _kart_ref(kart_id)
        if not kart:
            raise KartBulunamadi("Kart bulunamadı.")
        if not _operasyonda_gorunur_mu(kart):
            raise IsKuralHatasi("Bu kart operasyon ekranında aktif değil.")
        if kart["durum"] != PLANA_ALINDI:
            raise IsKuralHatasi("Yalnız PLANA ALINDI durumundaki kart DİZGİDE'ye alınabilir.")

        if adet in (None, ""):
            adet = kart["toplam_adet"]
        try:
            adet = int(adet)
        except (TypeError, ValueError) as exc:
            raise ValueError("Adet sayı olmalı.") from exc
        if adet < 1 or adet > kart["toplam_adet"]:
            raise IsKuralHatasi(f"Başlatılacak adet 1 ile {kart['toplam_adet']} arasında olmalı.")

        def islem():
            kart.update(
                durum=DIZGIDE,
                baslangic_adet=adet,
                baslama_zamani=kart.get("baslama_zamani") or simdi(),
                bitis_zamani=None,
                teslim_zamani=None,
                gerceklesen_teslim=None,
                operator=kullanici,
                aciklama=aciklama or kart.get("aciklama"),
                guncelleme=simdi(),
            )
            _loglar.append(
                _log_kaydi(
                    kullanici,
                    rol,
                    "DİZGİYE ALINDI",
                    kart.get("talep_no") or "",
                    kart.get("stok_no") or "",
                    adet,
                    aciklama or f"{adet} adet dizgiye alındı",
                )
            )
            return kart_gorunumu(kart)

        return _atomik_kart_islemi(islem)


def kart_bitir(kart_id, adet, kullanici, rol, aciklama=""):
    kart_id = _sayi(kart_id, -1)

    with _kilit:
        kart = _kart_ref(kart_id)
        if not kart:
            raise KartBulunamadi("Kart bulunamadı.")
        if kart.get("durum") != DIZGIDE:
            raise IsKuralHatasi("Tamamlanan adet yalnız DİZGİDE durumundaki karta girilebilir.")

        kalan = kart["toplam_adet"] - kart["tamamlanan_adet"]
        if kalan <= 0:
            raise IsKuralHatasi("Üretim adedi zaten tamamlandı. Kartı HAZIR durumuna alın.")

        try:
            adet = int(adet)
        except (TypeError, ValueError) as exc:
            raise ValueError("Adet sayı olmalı.") from exc
        if adet < 1 or adet > kalan:
            raise IsKuralHatasi(f"Adet 1 ile {kalan} arasında olmalı.")

        def islem():
            yeni_toplam = kart["tamamlanan_adet"] + adet
            uretim_bitti = yeni_toplam == kart["toplam_adet"]

            kart.update(
                tamamlanan_adet=yeni_toplam,
                bitis_zamani=simdi() if uretim_bitti else kart.get("bitis_zamani"),
                operator=kullanici,
                aciklama=aciklama or kart.get("aciklama"),
                guncelleme=simdi(),
            )
            _loglar.append(
                _log_kaydi(
                    kullanici,
                    rol,
                    "ÜRETİM ADEDİ TAMAMLANDI" if uretim_bitti else "KISMİ ÜRETİM",
                    kart.get("talep_no") or "",
                    kart.get("stok_no") or "",
                    adet,
                    aciklama or f"{yeni_toplam}/{kart['toplam_adet']} adet tamamlandı",
                )
            )

            mesaj = (
                "Üretim adedi tamamlandı. Kart HAZIR'a otomatik alınmadı."
                if uretim_bitti
                else f"{yeni_toplam}/{kart['toplam_adet']} adet tamamlandı."
            )
            return kart_gorunumu(kart), uretim_bitti, mesaj

        return _atomik_kart_islemi(islem)


def kart_hazirla(kart_id, kullanici, rol, aciklama=""):
    kart_id = _sayi(kart_id, -1)

    with _kilit:
        kart = _kart_ref(kart_id)
        if not kart:
            raise KartBulunamadi("Kart bulunamadı.")
        if kart.get("durum") != DIZGIDE:
            raise IsKuralHatasi("Yalnız DİZGİDE durumundaki kart HAZIR yapılabilir.")
        if kart.get("tamamlanan_adet", 0) != kart.get("toplam_adet", 0):
            raise IsKuralHatasi("Kart HAZIR yapılmadan önce üretim adedinin tamamı bitirilmelidir.")

        def islem():
            kart.update(
                durum=HAZIR,
                bitis_zamani=kart.get("bitis_zamani") or simdi(),
                operator=kullanici,
                aciklama=aciklama or kart.get("aciklama"),
                guncelleme=simdi(),
            )
            _loglar.append(
                _log_kaydi(
                    kullanici,
                    rol,
                    "HAZIR OLARAK İŞARETLENDİ",
                    kart.get("talep_no") or "",
                    kart.get("stok_no") or "",
                    kart.get("toplam_adet"),
                    aciklama or "Üretim tamamlandı; teslim alınmayı bekliyor.",
                )
            )
            return kart_gorunumu(kart)

        return _atomik_kart_islemi(islem)


def kart_teslim_et(kart_id, kullanici, rol, aciklama=""):
    kart_id = _sayi(kart_id, -1)

    with _kilit:
        kart = _kart_ref(kart_id)
        if not kart:
            raise KartBulunamadi("Kart bulunamadı.")
        if kart.get("durum") != HAZIR:
            raise IsKuralHatasi("Yalnız HAZIR durumundaki kart TESLİM EDİLDİ yapılabilir.")

        def islem():
            teslim_ani = simdi()
            kart.update(
                durum=TESLIM_EDILDI,
                gerceklesen_teslim=bugun(),
                teslim_zamani=teslim_ani,
                operator=kullanici,
                aciklama=aciklama or kart.get("aciklama"),
                guncelleme=teslim_ani,
            )
            _loglar.append(
                _log_kaydi(
                    kullanici,
                    rol,
                    "TESLİM EDİLDİ",
                    kart.get("talep_no") or "",
                    kart.get("stok_no") or "",
                    kart.get("toplam_adet"),
                    aciklama or f"Teslim tarihi: {bugun()}",
                )
            )
            return kart_gorunumu(kart)

        return _atomik_kart_islemi(islem)


def kart_not_guncelle(kart_id, aciklama, kullanici, rol):
    kart_id = _sayi(kart_id, -1)

    with _kilit:
        kart = _kart_ref(kart_id)
        if not kart:
            raise KartBulunamadi("Kart bulunamadı.")

        def islem():
            kart["aciklama"] = _temiz_metin(aciklama) or None
            kart["guncelleme"] = simdi()
            _loglar.append(
                _log_kaydi(
                    kullanici,
                    rol,
                    "NOT GÜNCELLENDİ",
                    kart.get("talep_no") or "",
                    kart.get("stok_no") or "",
                    detay=kart.get("aciklama") or "Not temizlendi",
                )
            )
            return kart_gorunumu(kart)

        return _atomik_kart_islemi(islem)


# ---------------------------------------------------------------------------
# Admin kart işlemleri
# ---------------------------------------------------------------------------

def _tarih_form_degeri(deger, alan_adi):
    if deger in (None, ""):
        return None
    sonuc = tarih_coz(deger)
    if not sonuc:
        raise IsKuralHatasi(f"{alan_adi} geçerli bir tarih olmalı.")
    return sonuc


def admin_kart_ekle(
    talep_no,
    stok_no,
    toplam_adet,
    kullanici,
    sira=None,
    talep_sahibi="",
    plan_hafta="",
    plan_baslama=None,
    plan_teslim=None,
    gerceklesen_teslim=None,
    pcb="",
    aciklama="",
):
    """Admin panelinden PLANA ALINDI durumunda manuel kart oluşturur."""
    global _kartlar, _loglar

    talep_no = _temiz_metin(talep_no)
    stok_no = _temiz_metin(stok_no)
    if not talep_no:
        raise IsKuralHatasi("Talep NO boş olamaz.")
    if not stok_no:
        raise IsKuralHatasi("Kart Stok No boş olamaz.")

    try:
        toplam = int(toplam_adet)
    except (TypeError, ValueError) as exc:
        raise ValueError("Toplam adet sayı olmalı.") from exc
    if toplam < 1:
        raise IsKuralHatasi("Toplam adet en az 1 olmalı.")

    if sira in (None, ""):
        sira_degeri = None
    else:
        try:
            sira_degeri = int(sira)
        except (TypeError, ValueError) as exc:
            raise ValueError("Sıra tam sayı olmalı.") from exc

    plan_baslama_iso = _tarih_form_degeri(plan_baslama, "Dizgi Başlama Tarihi")
    plan_teslim_iso = _tarih_form_degeri(plan_teslim, "Planlanan Teslim Tarihi")
    gerceklesen_iso = _tarih_form_degeri(gerceklesen_teslim, "Gerçekleşen Teslim Tarihi")

    if plan_baslama_iso and plan_teslim_iso and plan_baslama_iso > plan_teslim_iso:
        raise IsKuralHatasi("Dizgi Başlama Tarihi Planlanan Teslim Tarihinden sonra olamaz.")
    if gerceklesen_iso:
        raise IsKuralHatasi(
            "Yeni kart PLANA ALINDI durumunda başlar; Gerçekleşen Teslim Tarihi başlangıçta boş olmalı."
        )

    anahtar = f"{talep_no}|{stok_no}"

    with _kilit:
        if any(kart.get("anahtar") == anahtar for kart in _kartlar):
            raise IsKuralHatasi(f"Bu Talep NO + Kart Stok No zaten mevcut: {anahtar}")

        def islem():
            kart_id = max([_sayi(kart.get("id"), 0) for kart in _kartlar] or [0]) + 1
            kayit = {
                "id": kart_id,
                "sira": sira_degeri,
                "talep_no": talep_no,
                "stok_no": stok_no,
                "talep_sahibi": _temiz_metin(talep_sahibi) or None,
                "toplam_adet": toplam,
                "adet_metin": f"{toplam} ADET",
                "plan_hafta": _temiz_metin(plan_hafta) or None,
                "plan_baslama": plan_baslama_iso,
                "plan_teslim": plan_teslim_iso,
                "gerceklesen_teslim": None,
                "excel_durum": "MANUEL",
                "pcb": _temiz_metin(pcb) or None,
                "durum": PLANA_ALINDI,
                "baslangic_adet": 0,
                "tamamlanan_adet": 0,
                "baslama_zamani": None,
                "bitis_zamani": None,
                "teslim_zamani": None,
                "operator": None,
                "aciklama": _temiz_metin(aciklama) or None,
                "guncelleme": simdi(),
                "aktif": 1,
                "source_active": 1,
                "admin_gizli": 0,
                "kaynak": "MANUEL",
                "anahtar": anahtar,
            }
            _kartlar.append(kayit)
            _loglar.append(
                _log_kaydi(
                    kullanici,
                    "admin",
                    "MANUEL KART EKLENDİ",
                    talep_no,
                    stok_no,
                    toplam,
                    f"{toplam} adet · Durum: {PLANA_ALINDI}",
                )
            )
            return kart_gorunumu(kayit)

        return _atomik_kart_islemi(islem)


def admin_kart_duzenle(
    kart_id,
    durum,
    tamamlanan_adet,
    toplam_adet,
    aciklama,
    kullanici,
    plan_hafta=None,
    plan_baslama=None,
    plan_teslim=None,
    gerceklesen_teslim=None,
):
    """Admin workflow ve plan alanlarını kontrollü biçimde düzeltir."""
    kart_id = _sayi(kart_id, -1)

    with _kilit:
        kart = _kart_ref(kart_id)
        if not kart:
            raise KartBulunamadi("Kart bulunamadı.")

        yeni_durum = _durum_normalize(durum)
        if yeni_durum not in GECERLI_DURUMLAR:
            raise IsKuralHatasi("Durum PLANA ALINDI, DİZGİDE, HAZIR veya TESLİM EDİLDİ olmalı.")

        try:
            toplam = int(kart["toplam_adet"] if toplam_adet in (None, "") else toplam_adet)
            tamamlanan = int(
                kart["tamamlanan_adet"] if tamamlanan_adet in (None, "") else tamamlanan_adet
            )
        except (TypeError, ValueError) as exc:
            raise ValueError("Adetler sayı olmalı.") from exc

        if toplam < 1 or tamamlanan < 0 or tamamlanan > toplam:
            raise IsKuralHatasi("Adet değerleri tutarsız.")

        yeni_plan_hafta = (
            kart.get("plan_hafta")
            if plan_hafta is None
            else (_temiz_metin(plan_hafta) or None)
        )
        yeni_plan_baslama = (
            kart.get("plan_baslama")
            if plan_baslama is None
            else _tarih_form_degeri(plan_baslama, "Dizgi Başlama Tarihi")
        )
        yeni_plan_teslim = (
            kart.get("plan_teslim")
            if plan_teslim is None
            else _tarih_form_degeri(plan_teslim, "Planlanan Teslim Tarihi")
        )
        yeni_gerceklesen = (
            kart.get("gerceklesen_teslim")
            if gerceklesen_teslim is None
            else _tarih_form_degeri(gerceklesen_teslim, "Gerçekleşen Teslim Tarihi")
        )

        if yeni_plan_baslama and yeni_plan_teslim and yeni_plan_baslama > yeni_plan_teslim:
            raise IsKuralHatasi("Dizgi Başlama Tarihi Planlanan Teslim Tarihinden sonra olamaz.")

        onceki_durum = kart.get("durum") or "DURUMU EKSİK"

        def islem():
            nonlocal tamamlanan, yeni_gerceklesen

            baslama = kart.get("baslama_zamani")
            bitis = kart.get("bitis_zamani")
            teslim_zamani = kart.get("teslim_zamani")
            baslangic_adet = kart.get("baslangic_adet") or 0

            if yeni_durum == PLANA_ALINDI:
                tamamlanan = 0
                baslangic_adet = 0
                baslama = None
                bitis = None
                teslim_zamani = None
                yeni_gerceklesen = None

            elif yeni_durum == DIZGIDE:
                baslama = baslama or simdi()
                baslangic_adet = baslangic_adet or toplam
                bitis = bitis if tamamlanan == toplam else None
                teslim_zamani = None
                yeni_gerceklesen = None

            elif yeni_durum == HAZIR:
                tamamlanan = toplam
                baslama = baslama or simdi()
                bitis = bitis or simdi()
                teslim_zamani = None
                yeni_gerceklesen = None

            elif yeni_durum == TESLIM_EDILDI:
                tamamlanan = toplam
                baslama = baslama or simdi()
                bitis = bitis or simdi()
                yeni_gerceklesen = yeni_gerceklesen or bugun()
                teslim_zamani = teslim_zamani or simdi()

            kart.update(
                durum=yeni_durum,
                toplam_adet=toplam,
                tamamlanan_adet=tamamlanan,
                baslangic_adet=baslangic_adet,
                plan_hafta=yeni_plan_hafta,
                plan_baslama=yeni_plan_baslama,
                plan_teslim=yeni_plan_teslim,
                gerceklesen_teslim=yeni_gerceklesen,
                baslama_zamani=baslama,
                bitis_zamani=bitis,
                teslim_zamani=teslim_zamani,
                aciklama=_temiz_metin(aciklama) if aciklama is not None else kart.get("aciklama"),
                guncelleme=simdi(),
            )

            _loglar.append(
                _log_kaydi(
                    kullanici,
                    "admin",
                    "ADMİN DÜZENLEDİ",
                    kart.get("talep_no") or "",
                    kart.get("stok_no") or "",
                    tamamlanan,
                    f"{onceki_durum} → {yeni_durum} · {tamamlanan}/{toplam} adet",
                )
            )
            return kart_gorunumu(kart)

        return _atomik_kart_islemi(islem)


def admin_kart_gizle(kart_id, kullanici):
    kart_id = _sayi(kart_id, -1)

    with _kilit:
        kart = _kart_ref(kart_id)
        if not kart:
            raise KartBulunamadi("Kart bulunamadı.")

        def islem():
            kart["admin_gizli"] = 1
            kart["guncelleme"] = simdi()
            _loglar.append(
                _log_kaydi(
                    kullanici,
                    "admin",
                    "KART LİSTEDEN GİZLENDİ",
                    kart.get("talep_no") or "",
                    kart.get("stok_no") or "",
                    detay="Kart silinmedi; admin gizli olarak işaretlendi.",
                )
            )

        _atomik_kart_islemi(islem)


def admin_kart_geri_getir(kart_id, kullanici):
    kart_id = _sayi(kart_id, -1)

    with _kilit:
        kart = _kart_ref(kart_id)
        if not kart:
            raise KartBulunamadi("Kart bulunamadı.")
        if kart.get("admin_gizli", 0) != 1:
            raise IsKuralHatasi("Bu kart zaten görünür durumda.")

        def islem():
            kart["admin_gizli"] = 0
            kart["aktif"] = 1
            kart["guncelleme"] = simdi()
            _loglar.append(
                _log_kaydi(
                    kullanici,
                    "admin",
                    "GİZLENEN KART GERİ GETİRİLDİ",
                    kart.get("talep_no") or "",
                    kart.get("stok_no") or "",
                    detay="Kart tekrar yönetim/operasyon listelerine alındı.",
                )
            )
            return kart_gorunumu(kart)

        return _atomik_kart_islemi(islem)


# ---------------------------------------------------------------------------
# Yedekten geri yükleme
# ---------------------------------------------------------------------------

def yedekten_geri_yukle(yedek_adi, kullanici):
    global _kartlar, _loglar

    with _kilit:
        aday = _yedek_kart_dosyasi_bul(yedek_adi)
        yeni_kartlar = _oku(aday, KART_ALANLARI, ZORUNLU_KART_ALANLARI)
        yeni_kartlar = [_kart_normalize(kart) for kart in yeni_kartlar]
        if not yeni_kartlar:
            raise VeriDogrulamaHatasi("Seçilen yedekte hiç kart bulunmuyor.")
        _kart_listesi_dogrula(yeni_kartlar)

        koruma_yedegi = anlik_yedek("geri_yukleme_oncesi")
        eski_kartlar = copy.deepcopy(_kartlar)
        eski_loglar = copy.deepcopy(_loglar)

        try:
            _kartlar = yeni_kartlar
            _loglar.append(
                _log_kaydi(
                    kullanici,
                    "admin",
                    "YEDEKTEN GERİ YÜKLENDİ",
                    detay=(
                        f"Yedek: {yedek_adi} · {len(_kartlar)} kart · "
                        f"koruma: {os.path.basename(koruma_yedegi)}"
                    ),
                )
            )
            _coklu_yaz(
                [
                    (KARTLAR_DOSYA, KART_ALANLARI, _kartlar, "Kartlar"),
                    (LOG_DOSYA, LOG_ALANLARI, _loglar, "İşlem Logu"),
                ]
            )
        except Exception:
            _kartlar = eski_kartlar
            _loglar = eski_loglar
            raise

        return {
            "kart": len(_kartlar),
            "yedek": yedek_adi,
            "koruma_yedegi": koruma_yedegi,
        }


# ---------------------------------------------------------------------------
# Excel import commit
# ---------------------------------------------------------------------------

def excel_import_uygula(dosya_adi, kullanici, satirlar, uyari_sayisi=0):
    """Tamamen parse/validate edilmiş kaynak satırlarını tek transaction-benzeri blokta uygular."""
    global _kartlar, _loglar, _yuklemeler

    with _kilit:
        eski_kartlar = copy.deepcopy(_kartlar)
        eski_loglar = copy.deepcopy(_loglar)
        eski_yuklemeler = copy.deepcopy(_yuklemeler)

        try:
            yedek_klasoru = anlik_yedek("import_oncesi")
            mevcut_harita = {kart.get("anahtar"): kart for kart in _kartlar}
            gorulen = set()
            yeni = 0
            guncellenen = 0
            workflow_korundu = 0
            pasife_listesi = []
            sonraki_id = max([_sayi(kart.get("id"), 0) for kart in _kartlar] or [0]) + 1

            for satir in satirlar:
                anahtar = satir["anahtar"]
                gorulen.add(anahtar)
                plan = satir["plan"]
                mevcut = mevcut_harita.get(anahtar)

                if mevcut:
                    yeni_toplam = int(plan["toplam_adet"])
                    tamamlanan = int(mevcut.get("tamamlanan_adet") or 0)
                    if yeni_toplam < tamamlanan:
                        raise IsKuralHatasi(
                            f"{anahtar}: Excel toplam adedi ({yeni_toplam}), sistemde tamamlanan "
                            f"adetten ({tamamlanan}) küçük olamaz. Import iptal edildi."
                        )
                    if mevcut.get("durum") in (HAZIR, TESLIM_EDILDI) and yeni_toplam != tamamlanan:
                        raise IsKuralHatasi(
                            f"{anahtar}: Kart {mevcut['durum']} ve {tamamlanan} adet tamamlanmış. "
                            f"Yeni Excel toplam adedi {yeni_toplam}; admin kontrolü gerekir."
                        )

                    workflow = {
                        "durum": mevcut.get("durum"),
                        "baslangic_adet": mevcut.get("baslangic_adet"),
                        "tamamlanan_adet": tamamlanan,
                        "baslama_zamani": mevcut.get("baslama_zamani"),
                        "bitis_zamani": mevcut.get("bitis_zamani"),
                        "teslim_zamani": mevcut.get("teslim_zamani"),
                        "operator": mevcut.get("operator"),
                        "aciklama": mevcut.get("aciklama"),
                        "gerceklesen_teslim": mevcut.get("gerceklesen_teslim"),
                    }

                    mevcut.update(plan)
                    mevcut.update(workflow)
                    mevcut["source_active"] = 1
                    mevcut["aktif"] = 1
                    mevcut["kaynak"] = "EXCEL"
                    mevcut["guncelleme"] = simdi()
                    _kart_dogrula(mevcut)
                    guncellenen += 1
                    workflow_korundu += 1
                    continue

                toplam = int(plan["toplam_adet"])
                durum = satir.get("ilk_durum")
                gerceklesen = satir.get("gerceklesen_teslim")

                tamamlanan = toplam if durum in (HAZIR, TESLIM_EDILDI) else 0
                baslangic_adet = (
                    toplam
                    if durum in (DIZGIDE, HAZIR, TESLIM_EDILDI)
                    else 0
                )

                baslama = None
                bitis = None
                teslim_zamani = None

                kayit = {
                    "id": sonraki_id,
                    "anahtar": anahtar,
                    "talep_no": satir["talep_no"],
                    "stok_no": satir["stok_no"],
                    "gerceklesen_teslim": (
                        gerceklesen
                        if durum == TESLIM_EDILDI
                        else None
                    ),
                    "durum": durum,
                    "baslangic_adet": baslangic_adet,
                    "tamamlanan_adet": tamamlanan,
                    "baslama_zamani": None,
                    "bitis_zamani": None,
                    "teslim_zamani": None,
                    "operator": "Excel" if durum in GECERLI_DURUMLAR else None,
                    "aciklama": None,
                    "guncelleme": simdi(),
                    "aktif": 1,
                    "source_active": 1,
                    "admin_gizli": 0,
                    "kaynak": "EXCEL",
                }
                kayit.update(plan)
                _kart_dogrula(kayit)
                _kartlar.append(kayit)
                mevcut_harita[anahtar] = kayit
                sonraki_id += 1
                yeni += 1

            pasife_alinan = 0
            for kart in _kartlar:
                if kart.get("kaynak") != "EXCEL" or kart.get("anahtar") in gorulen:
                    continue
                if kart.get("source_active", 1) == 1:
                    kart["source_active"] = 0
                    kart["guncelleme"] = simdi()
                    pasife_alinan += 1
                    pasife_listesi.append(
                        f"{kart.get('talep_no') or ''}|{kart.get('stok_no') or ''}"
                    )

            _kart_listesi_dogrula(_kartlar)

            _yuklemeler.append(
                {
                    "zaman": simdi(),
                    "kullanici": kullanici,
                    "dosya": dosya_adi,
                    "satir": len(satirlar),
                    "yeni": yeni,
                    "guncellenen": guncellenen,
                    "pasife_alinan": pasife_alinan,
                    "uyari": uyari_sayisi,
                }
            )
            _loglar.append(
                _log_kaydi(
                    kullanici,
                    "admin",
                    "EXCEL YÜKLENDİ",
                    detay=(
                        f"{len(satirlar)} satır · {yeni} yeni · {guncellenen} güncellendi · "
                        f"{workflow_korundu} workflow korundu · {pasife_alinan} kaynakta yok · "
                        f"{uyari_sayisi} uyarı · yedek={os.path.basename(yedek_klasoru)}"
                    ),
                )
            )

            _gunluk_yedek(KARTLAR_DOSYA)
            _coklu_yaz(
                [
                    (KARTLAR_DOSYA, KART_ALANLARI, _kartlar, "Kartlar"),
                    (LOG_DOSYA, LOG_ALANLARI, _loglar, "İşlem Logu"),
                    (YUKLEME_DOSYA, YUKLEME_ALANLARI, _yuklemeler, "Yüklemeler"),
                ]
            )

            return {
                "satir": len(satirlar),
                "yeni": yeni,
                "guncellenen": guncellenen,
                "pasife_alinan": pasife_alinan,
                "pasife_listesi": pasife_listesi[:20],
                "workflow_korundu": workflow_korundu,
                "uyari": uyari_sayisi,
                "yedek": yedek_klasoru,
            }

        except Exception:
            _kartlar = eski_kartlar
            _loglar = eski_loglar
            _yuklemeler = eski_yuklemeler
            raise


# ---------------------------------------------------------------------------
# Legacy küçük yardımcılar
# ---------------------------------------------------------------------------

def kart_guncelle(kart_id, **alanlar):
    global _kartlar

    kart_id = _sayi(kart_id, -1)

    with _kilit:
        index = next(
            (
                i
                for i, kart in enumerate(_kartlar)
                if kart.get("id") == kart_id
            ),
            None,
        )

        if index is None:
            return None

        eski = copy.deepcopy(_kartlar)

        try:
            yeni = copy.deepcopy(_kartlar[index])
            yeni.update(alanlar)
            yeni["guncelleme"] = simdi()

            if "talep_no" in alanlar or "stok_no" in alanlar:
                yeni["anahtar"] = (
                    f"{_temiz_metin(yeni.get('talep_no'))}|"
                    f"{_temiz_metin(yeni.get('stok_no'))}"
                )

            yeni = _kart_normalize(yeni)
            _kartlar[index] = yeni

            _kart_listesi_dogrula(_kartlar)
            _kartlari_kaydet()

            return kart_gorunumu(yeni)

        except Exception:
            _kartlar = eski
            raise


def toplu_kaydet(yeni_kartlar=(), degisen=True):
    global _kartlar

    with _kilit:
        eski = copy.deepcopy(_kartlar)
        try:
            _kartlar.extend(_kart_normalize(kart) for kart in yeni_kartlar)
            _kart_listesi_dogrula(_kartlar)
            if degisen:
                _kartlari_kaydet()
        except Exception:
            _kartlar = eski
            raise
