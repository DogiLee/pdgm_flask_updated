"""PDGM kaynak Excel importu ve rapor üretimi.

Kaynak import kuralları:
- Yalnız MAKİNE sayfası kullanılır.
- Hidden kolonlar alınmaz; hidden satırlar korunur.
- Microsoft Excel COM ile values-only snapshot oluşturulur.
- External link güncellemesi ve full recalculation yapılmaz.
- Talep NO + Kart Stok No olmayan satırlar kart sayılmaz.
- Workflow yalnız şu dört durumdan oluşur:
    PLANA ALINDI, DİZGİDE, HAZIR, TESLİM EDİLDİ
- DURUM boş veya farklı bir değer ise kart kaybolmaz; durum None olur ve
  operasyon ekranında gösterilmez. Admin Yönetim ekranında düzeltebilir.
"""

from __future__ import annotations

import gc
import io
import os
import re
import tempfile
import threading
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    import pythoncom
    import win32com.client
except ImportError:
    pythoncom = None
    win32com = None

import depo

_import_kilidi = threading.Lock()


class ExcelAktarimHatasi(Exception):
    pass


KAYNAK_SAYFA_ADI = "MAKİNE"
MSO_AUTOMATION_SECURITY_FORCE_DISABLE = 3


TURKCE_HARFLER = str.maketrans(
    {
        "ç": "c",
        "Ç": "C",
        "ğ": "g",
        "Ğ": "G",
        "ı": "i",
        "İ": "I",
        "ö": "o",
        "Ö": "O",
        "ş": "s",
        "Ş": "S",
        "ü": "u",
        "Ü": "U",
        "â": "a",
        "Â": "A",
    }
)


def _sadelestir(deger):
    metin = str(deger or "").translate(TURKCE_HARFLER).upper()
    metin = metin.replace(".", " ").replace("_", " ").replace("\xa0", " ")
    return re.sub(r"\s+", " ", metin).strip()


BASLIK_ESLESME = {
    _sadelestir(baslik): alan
    for baslik, alan in {
        "NO": "sira",
        "Talep NO": "talep_no",
        "Talep Sahibi": "talep_sahibi",
        "Kart Stok No": "stok_no",
        "Kart Üretim Adet": "adet_metin",
        "Planlanan Başlangıç T.": "plan_hafta",
        "Dizgi Başlama Tarihi": "plan_baslama",
        "Planlanan Teslim T.": "plan_teslim",
        "Gerçekleşen Teslim T.": "gerceklesen_teslim",
        "DURUM": "excel_durum",
        "PCB": "pcb",
        # Küçük format toleransı
        "Sıra": "sira",
        "Üretim Adet": "adet_metin",
        "Adet": "adet_metin",
    }.items()
}

DURUM_ESLESME = {
    "PLANA ALINDI": depo.PLANA_ALINDI,
    "DIZGIDE": depo.DIZGIDE,
    "HAZIR": depo.HAZIR,
    "TESLIM EDILDI": depo.TESLIM_EDILDI,
}


# ---------------------------------------------------------------------------
# COM snapshot
# ---------------------------------------------------------------------------

def excel_deger_snapshot_olustur(dosya_yolu):
    """MAKİNE sheet'indeki visible kolonları values-only geçici xlsx'e kopyalar."""
    if pythoncom is None or win32com is None:
        raise ExcelAktarimHatasi(
            "Excel COM aktarımı yalnızca Windows + Microsoft Excel ortamında çalışır "
            "(pywin32 gerekli)."
        )

    kaynak = Path(dosya_yolu).resolve()
    temp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    temp.close()
    hedef = Path(temp.name).resolve()

    pythoncom.CoInitialize()
    excel = None
    kaynak_wb = None
    hedef_wb = None
    kaynak_ws = None
    hedef_ws = None
    used = None
    kaynak_aralik = None
    hedef_aralik = None
    basarili = False

    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        excel.EnableEvents = False
        excel.AutomationSecurity = MSO_AUTOMATION_SECURITY_FORCE_DISABLE

        kaynak_wb = excel.Workbooks.Open(
            str(kaynak),
            UpdateLinks=0,
            ReadOnly=True,
            IgnoreReadOnlyRecommended=True,
            AddToMru=False,
        )

        kaynak_ws = next(
            (
                ws
                for ws in kaynak_wb.Worksheets
                if _sadelestir(ws.Name) == _sadelestir(KAYNAK_SAYFA_ADI)
            ),
            None,
        )
        if kaynak_ws is None:
            sayfalar = [ws.Name for ws in kaynak_wb.Worksheets]
            raise ExcelAktarimHatasi(
                f"'{KAYNAK_SAYFA_ADI}' sayfası bulunamadı. "
                f"Dosyadaki sayfalar: {', '.join(sayfalar)}"
            )

        hedef_wb = excel.Workbooks.Add()
        while hedef_wb.Worksheets.Count > 1:
            hedef_wb.Worksheets(hedef_wb.Worksheets.Count).Delete()

        hedef_ws = hedef_wb.Worksheets(1)
        hedef_ws.Name = KAYNAK_SAYFA_ADI

        used = kaynak_ws.UsedRange
        ilk_satir = used.Row
        son_satir = used.Row + used.Rows.Count - 1
        ilk_sutun = used.Column
        son_sutun = used.Column + used.Columns.Count - 1
        hedef_sutun = 1

        for kaynak_sutun in range(ilk_sutun, son_sutun + 1):
            if bool(kaynak_ws.Columns(kaynak_sutun).Hidden):
                continue

            kaynak_aralik = kaynak_ws.Range(
                kaynak_ws.Cells(ilk_satir, kaynak_sutun),
                kaynak_ws.Cells(son_satir, kaynak_sutun),
            )
            hedef_aralik = hedef_ws.Range(
                hedef_ws.Cells(ilk_satir, hedef_sutun),
                hedef_ws.Cells(son_satir, hedef_sutun),
            )
            hedef_aralik.Value2 = kaynak_aralik.Value2
            hedef_sutun += 1

        hedef_wb.SaveAs(str(hedef), FileFormat=51)
        basarili = True
        return str(hedef)

    finally:
        if hedef_wb is not None:
            try:
                hedef_wb.Close(SaveChanges=False)
            except Exception:
                pass
        if kaynak_wb is not None:
            try:
                kaynak_wb.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass

        hedef_aralik = None
        kaynak_aralik = None
        used = None
        hedef_ws = None
        kaynak_ws = None
        hedef_wb = None
        kaynak_wb = None
        excel = None
        gc.collect()

        pythoncom.CoUninitialize()

        if not basarili:
            try:
                os.remove(str(hedef))
            except OSError:
                pass


# ---------------------------------------------------------------------------
# Parser helpers
# ---------------------------------------------------------------------------

def durum_coz(deger):
    """Kaynak durumunu dört gerçek workflow durumundan birine çevirir."""
    temiz = _sadelestir(deger)
    if not temiz:
        return None, True
    durum = DURUM_ESLESME.get(temiz)
    return durum, durum is None


def adet_coz(deger):
    """Üretim adedini pozitif tam sayıya çevirir.

    Kabul:
    - 400
    - 400.0
    - "400 ADET"
    - "1.500 ADET"
    - "1,500 ADET"
    - "1 500 ADET"

    Red:
    - 400.5
    - "400.5 ADET"
    - "400,5 ADET"
    - "-5 ADET"
    - "400 / 500 ADET"
    """
    if deger is None or str(deger).strip() == "":
        raise ExcelAktarimHatasi("Üretim adedi boş olamaz.")

    if isinstance(deger, bool):
        raise ExcelAktarimHatasi("Üretim adedi sayı olmalı.")

    if isinstance(deger, (int, float)):
        if isinstance(deger, float) and not deger.is_integer():
            raise ExcelAktarimHatasi(
                f"Üretim adedi tam sayı olmalı: {deger}"
            )
        adet = int(deger)

    else:
        metin = str(deger).strip()

        eslesmeler = [
            eslesme.strip()
            for eslesme in re.findall(
                r"[-+]?\d[\d.,\s]*",
                metin,
            )
            if eslesme.strip()
        ]

        if len(eslesmeler) != 1:
            raise ExcelAktarimHatasi(
                f"Üretim adedi tek bir sayı içermeli: '{metin}'"
            )

        token = re.sub(r"\s+", "", eslesmeler[0])

        if token.startswith(("+", "-")):
            raise ExcelAktarimHatasi(
                f"Üretim adedi pozitif tam sayı olmalı: '{metin}'"
            )

        if token.isdigit():
            adet = int(token)

        elif re.fullmatch(
            r"\d{1,3}(?:[.,]\d{3})+",
            token,
        ):
            adet = int(re.sub(r"[.,]", "", token))

        else:
            raise ExcelAktarimHatasi(
                f"Üretim adedi tam sayı olmalı: '{metin}'"
            )

    if adet < 1:
        raise ExcelAktarimHatasi(
            f"Üretim adedi en az 1 olmalı: {deger!r}"
        )

    return adet


def _baslik_satiri_bul(ws):
    for satir_no, satir in enumerate(
        ws.iter_rows(min_row=1, max_row=10, values_only=True),
        start=1,
    ):
        temiz = [_sadelestir(hucre) for hucre in satir]
        if "TALEP NO" in temiz or "KART STOK NO" in temiz:
            return satir_no, temiz
    return None, None


def _hucre_al(satir, kolonlar, alan):
    index = kolonlar.get(alan)
    if index is None or index >= len(satir):
        return None
    deger = satir[index]
    return None if deger == "" else deger


def _tarih_coz_ve_dogrula(deger, alan_adi, excel_satir_no):
    sonuc = depo.tarih_coz(deger)
    if deger not in (None, "") and not sonuc:
        raise ExcelAktarimHatasi(
            f"Excel satır {excel_satir_no}: {alan_adi} okunamadı. Gelen değer: {deger!r}"
        )
    return sonuc


def _sira_coz(deger):
    if deger in (None, ""):
        return None, False
    try:
        return int(float(deger)), False
    except (TypeError, ValueError):
        return None, True


# ---------------------------------------------------------------------------
# Import
# ---------------------------------------------------------------------------

def excelden_aktar(dosya_yolu, kullanici):
    with _import_kilidi:
        return _excelden_aktar(dosya_yolu, kullanici)


def _excelden_aktar(dosya_yolu, kullanici):
    snapshot_yolu = None
    wb = None

    try:
        snapshot_yolu = excel_deger_snapshot_olustur(dosya_yolu)
        wb = openpyxl.load_workbook(snapshot_yolu, data_only=True, read_only=True)
    except ExcelAktarimHatasi:
        raise
    except Exception as exc:
        raise ExcelAktarimHatasi(
            f"Dosya geçerli bir Excel çalışma kitabı değil: {exc}"
        ) from exc

    try:
        if KAYNAK_SAYFA_ADI not in wb.sheetnames:
            raise ExcelAktarimHatasi(
                f"'{KAYNAK_SAYFA_ADI}' sayfası bulunamadı. "
                f"Dosyadaki sayfalar: {', '.join(wb.sheetnames)}"
            )

        ws = wb[KAYNAK_SAYFA_ADI]
        baslik_no, basliklar = _baslik_satiri_bul(ws)
        if not baslik_no:
            raise ExcelAktarimHatasi(
                "Başlık satırı bulunamadı. Dosyada 'Talep NO' veya 'Kart Stok No' sütunu olmalı."
            )

        kolonlar = {}
        for index, baslik in enumerate(basliklar):
            alan = BASLIK_ESLESME.get(baslik)
            if not alan:
                continue
            if alan in kolonlar:
                raise ExcelAktarimHatasi(
                    f"'{baslik}' için birden fazla görünür Excel sütunu bulundu."
                )
            kolonlar[alan] = index

        for gerekli in ("talep_no", "stok_no", "adet_metin"):
            if gerekli not in kolonlar:
                ad = {
                    "talep_no": "Talep NO",
                    "stok_no": "Kart Stok No",
                    "adet_metin": "Kart Üretim Adet",
                }[gerekli]
                raise ExcelAktarimHatasi(f"Zorunlu '{ad}' sütunu bulunamadı.")

        parsed = []
        gorulen_anahtarlar = set()
        uyari_sayisi = 0

        for excel_satir_no, satir in enumerate(
            ws.iter_rows(min_row=baslik_no + 1, values_only=True),
            start=baslik_no + 1,
        ):
            if not any(hucre not in (None, "") for hucre in satir):
                continue

            talep_no = str(_hucre_al(satir, kolonlar, "talep_no") or "").strip()
            stok_no = str(_hucre_al(satir, kolonlar, "stok_no") or "").strip()
            if not talep_no or not stok_no:
                continue

            anahtar = f"{talep_no}|{stok_no}"
            if anahtar in gorulen_anahtarlar:
                raise ExcelAktarimHatasi(
                    f"Excel satır {excel_satir_no}: Talep NO + Kart Stok No tekrar ediyor ({anahtar})."
                )
            gorulen_anahtarlar.add(anahtar)

            try:
                toplam_adet = adet_coz(_hucre_al(satir, kolonlar, "adet_metin"))
            except ExcelAktarimHatasi as exc:
                raise ExcelAktarimHatasi(f"Excel satır {excel_satir_no}: {exc}") from exc

            plan_baslama_ham = _hucre_al(satir, kolonlar, "plan_baslama")
            plan_teslim_ham = _hucre_al(satir, kolonlar, "plan_teslim")
            gerceklesen_ham = _hucre_al(satir, kolonlar, "gerceklesen_teslim")

            plan_baslama = _tarih_coz_ve_dogrula(
                plan_baslama_ham, "Dizgi Başlama Tarihi", excel_satir_no
            )
            plan_teslim = _tarih_coz_ve_dogrula(
                plan_teslim_ham, "Planlanan Teslim Tarihi", excel_satir_no
            )
            gerceklesen = _tarih_coz_ve_dogrula(
                gerceklesen_ham, "Gerçekleşen Teslim Tarihi", excel_satir_no
            )

            if plan_baslama and plan_teslim and plan_baslama > plan_teslim:
                raise ExcelAktarimHatasi(
                    f"Excel satır {excel_satir_no}: Dizgi Başlama Tarihi ({plan_baslama}) "
                    f"Planlanan Teslim Tarihinden ({plan_teslim}) sonra olamaz."
                )

            excel_durum_raw = _hucre_al(satir, kolonlar, "excel_durum")
            ilk_durum, durum_uyarisi = durum_coz(excel_durum_raw)
            if durum_uyarisi:
                uyari_sayisi += 1

            if (ilk_durum == depo.TESLIM_EDILDI and not gerceklesen):
                ilk_durum = None
                uyari_sayisi +=1

            elif (ilk_durum != depo.TESLIM_EDILDI and gerceklesen):
                uyari_sayisi += 1

            sira, sira_uyarisi = _sira_coz(_hucre_al(satir, kolonlar, "sira"))
            if sira_uyarisi:
                uyari_sayisi += 1

            plan = {
                "sira": sira,
                "talep_sahibi": str(_hucre_al(satir, kolonlar, "talep_sahibi") or "").strip(),
                "toplam_adet": toplam_adet,
                "adet_metin": str(_hucre_al(satir, kolonlar, "adet_metin") or "").strip(),
                "plan_hafta": str(_hucre_al(satir, kolonlar, "plan_hafta") or "").strip(),
                "plan_baslama": plan_baslama,
                "plan_teslim": plan_teslim,
                "excel_durum": str(excel_durum_raw or "").strip(),
                "pcb": str(_hucre_al(satir, kolonlar, "pcb") or "").strip(),
            }

            parsed.append(
                {
                    "anahtar": anahtar,
                    "talep_no": talep_no,
                    "stok_no": stok_no,
                    "plan": plan,
                    "gerceklesen_teslim": gerceklesen,
                    "ilk_durum": ilk_durum,
                }
            )

        if not parsed:
            raise ExcelAktarimHatasi("Excel'de işlenecek kart satırı bulunamadı.")

    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
        if snapshot_yolu:
            try:
                os.remove(snapshot_yolu)
            except OSError:
                pass

    return depo.excel_import_uygula(
        dosya_adi=os.path.basename(dosya_yolu),
        kullanici=kullanici,
        satirlar=parsed,
        uyari_sayisi=uyari_sayisi,
    )


# ---------------------------------------------------------------------------
# Rapor üretimi
# ---------------------------------------------------------------------------

BASLIK_DOLGU = PatternFill("solid", fgColor="0F2027")
BASLIK_YAZI = Font(name="Arial", bold=True, color="FFFFFF", size=11)
GOVDE_YAZI = Font(name="Arial", size=10)

def _excel_hucre_yaz(hucre, deger):
    hucre.value = deger

    if isinstance(deger, str) and deger.startswith("="):
        hucre.data_type = "s"

def _sayfa_yaz(ws, basliklar, satirlar):
    ws.append(basliklar)
    for hucre in ws[1]:
        hucre.fill = BASLIK_DOLGU
        hucre.font = BASLIK_YAZI
        hucre.alignment = Alignment(horizontal="center", vertical="center")

    for satir in satirlar:
        satir_no = ws.max_row + 1

        for sutun_no, deger in enumerate(satir,start=1):
            _excel_hucre_yaz(
                ws.cell(row=satir_no,column=sutun_no,),deger)


    for sutun in range(1, len(basliklar) + 1):
        en = max(
            [len(str(basliklar[sutun - 1])), 10]
            + [len(str(satir[sutun - 1])) for satir in satirlar[:400]]
        )
        ws.column_dimensions[get_column_letter(sutun)].width = min(40, en + 3)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for satir in ws.iter_rows(min_row=2):
        for hucre in satir:
            hucre.font = GOVDE_YAZI


def calisma_kitabi_uret(kartlar, loglar, ozet):
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Kart Durumları"
    _sayfa_yaz(
        ws,
        [
            "ID",
            "Sıra",
            "Talep NO",
            "Talep Sahibi",
            "Kart Stok No",
            "Toplam Adet",
            "Tamamlanan",
            "Kalan",
            "Durum",
            "Kaynak Durumu",
            "Değerlendirme",
            "Sapma (gün)",
            "Plan Başlangıç",
            "Plan Teslim",
            "Üretim Başlangıç",
            "Üretim Bitiş",
            "Teslim Zamanı",
            "Gerçekleşen Teslim",
            "Operatör",
            "Not",
            "PCB",
            "Kaynakta Aktif",
            "Admin Gizli",
        ],
        [
            [
                k.get("id"),
                k.get("sira"),
                k.get("talep_no"),
                k.get("talep_sahibi"),
                k.get("stok_no"),
                k.get("toplam_adet"),
                k.get("tamamlanan_adet"),
                k.get("kalan_adet"),
                k.get("durum") or "DURUMU EKSİK",
                k.get("excel_durum") or "",
                k.get("rozet"),
                k.get("sapma") if k.get("sapma") is not None else "",
                k.get("plan_baslama") or "",
                k.get("plan_teslim") or "",
                k.get("baslama_zamani") or "",
                k.get("bitis_zamani") or "",
                k.get("teslim_zamani") or "",
                k.get("gerceklesen_teslim") or "",
                k.get("operator") or "",
                k.get("aciklama") or "",
                k.get("pcb") or "",
                k.get("source_active", 1),
                k.get("admin_gizli", 0),
            ]
            for k in kartlar
        ],
    )

    ws2 = wb.create_sheet("İşlem Logu")
    _sayfa_yaz(
        ws2,
        ["Zaman", "Kullanıcı", "Rol", "İşlem", "Talep NO", "Kart Stok No", "Adet", "Detay"],
        [
            [
                l.get("zaman"),
                l.get("kullanici"),
                l.get("rol"),
                l.get("islem"),
                l.get("talep_no") or "",
                l.get("stok_no") or "",
                l.get("adet") if l.get("adet") is not None else "",
                l.get("detay") or "",
            ]
            for l in loglar
        ],
    )

    ws3 = wb.create_sheet("Özet")
    _sayfa_yaz(ws3, ["Başlık", "Değer"], ozet)
    return wb


def kitap_baytlari(wb):
    tampon = io.BytesIO()
    wb.save(tampon)
    tampon.seek(0)
    return tampon


def dosya_adi(on_ek):
    return f"{on_ek}_{datetime.now():%Y%m%d_%H%M}.xlsx"
